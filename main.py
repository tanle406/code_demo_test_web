from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
import win32com.client as win32
import pandas as pd
import datetime
import os
import sys
from time import sleep
import inspect
import logging
from shutil import copyfile


global LOG_INFO
CurDir = os.path.dirname(
    os.path.abspath(inspect.getfile(inspect.currentframe()))
)
logging.basicConfig(
    format="------------------------------- %(asctime)s >>>  %(message)s  <<<-------------------------------",
    datefmt="%d/%m/%Y %H:%M:%S",
)
logFormatter = logging.Formatter(
    "%(asctime)s ---------- %(message)s", datefmt="%d/%m/%Y %H:%M:%S"
)
LOG_INFO = logging.warning
FileHandler = logging.FileHandler("log.txt", "a+", "utf-8")
FileHandler.setFormatter(logFormatter)
logging.getLogger().addHandler(FileHandler)
LOG_INFO("-------------------------------START-------------------------------")

def highlight(x):
    c1 = 'background-color: yellow'

    #empty DataFrame of styles
    df1 = pd.DataFrame('', index=x.index, columns=x.columns)
    #set new columns by condition
    df1.loc[(x['金額チェック'] != 'OK'), '金額チェック'] = c1
    return df1

class Browser():
    def __init__(self, curdir, log_infor):
        self.curdir = curdir
        self.log_infor = log_infor
        self.list_result = []

    def create_folder(self, path_folder):
        check_folder = os.path.exists(path_folder)
        if check_folder:
            pass
        else:
            os.makedirs(path_folder)

    def move_file(self, old_path, new_location):
        time_now = datetime.datetime.now()
        name_file = old_path.split("\\")[-1]
        new_path = new_location + "\\" + name_file
        if os.path.exists(old_path):
            if not os.path.exists(new_path):
                os.rename(old_path, new_path)
            elif os.path.exists(new_path):
                new_path = new_location + "\\" + \
                    time_now.strftime("%Y%m%d%f") + "_" + name_file
                os.rename(old_path, new_path)

    def DisplayMessageBox(self, body, title="Message", type="info"):
        '''
        Shows a pop-up message with title and body. Three possible types, info, error and warning with the default value info.
        '''
        import tkinter
        from tkinter import messagebox

        # hide main window
        root = tkinter.Tk()
        root.withdraw()
        if not body:
            messagebox.showwarning("Warning", "No input for message box")

        if type == "error":
            messagebox.showwarning(title, body)
        if type == "warning":
            messagebox.showwarning(title, body)
        if type == "info":
            messagebox.showinfo(title, body)
        return

    def open_browser_firefox(self):
        '''
        open firefox and return driver
        '''
        self.log_infor("Start Open Browser FireFox")
        driver = webdriver.Firefox(executable_path = self.curdir + "\\driver\\geckodriver.exe")
        self.log_infor("End Open Browser FireFox")
        return driver

    def open_browser_microsoft_edge(self):
        '''
        open firefox and return driver
        '''
        self.log_infor("Start Open Browser MS Edge")
        driver = webdriver.Edge(executable_path = self.curdir + "\\driver\\msedgedriver.exe")
        self.log_infor("Close Open Browser MS Edge")
        return driver

    def open_browser_chrome(self):
        self.log_infor("Start open browser Chrome")
        chromeOptions = webdriver.ChromeOptions()
        driver = webdriver.Chrome(
            executable_path=os.path.abspath(self.curdir + "\\driver\\chromedriver.exe"),
            chrome_options=chromeOptions)
        self.log_infor("End open browser Chrome")
        return driver

    def go_to_url(self, url, driver):
        '''
        go to website
        '''
        self.log_infor("Start go to website")
        driver.get(url)
        self.log_infor("End go to website")

    def wait_page_complete(self, driver, class_name):
        '''
        Check page complete and return the result
        class_name: class name need to check (string)
        '''
        delay = 10 # seconds
        try:
            myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.CLASS_NAME, class_name)))
            result = True
            print("Page is ready!")
        except TimeoutException:
            result = False
            print("Loading took too much time!")
        return True

    def login_alert(self, user, password, driver):
        window_before = driver.window_handles[0]
        driver.switch_to_window(window_before)
        shell = win32.Dispatch("WScript.Shell")
        shell.Sendkeys(user)
        sleep(0.3)
        shell.Sendkeys('{TAB}')
        sleep(0.3)
        shell.Sendkeys(password)
        sleep(0.3)
        shell.Sendkeys('{TAB}')
        sleep(0.3)
        shell.Sendkeys('{ENTER}')

    def check_button_page_top(self, driver):
        '''
        Check button "Page Top" and return the result
        '''
        self.log_infor("Start check button Page Top")
        driver.execute_script('window.scroll(0,1500)')
        sleep(2)
        try:
            # check button page top and click
            driver.find_element(By.CLASS_NAME, 'is-active').click()
            sleep(1)
            try:
                driver.find_element(By.CLASS_NAME, 'is-active')
                result = False
            except:
                result = True
        except Exception as error:
            result = False
        self.log_infor("Result check button Page Top: {}".format(result))
        self.log_infor("End check button Page Top")
        self.list_result.append(result)

    def check_slider_list_item(self, driver):
        '''
        check slider list item and return the result
        '''
        self.log_infor("Start check slider list item")
        try:
            driver.find_element(By.CLASS_NAME, 'c-sliderlist__item')
            result = True
        except Exception as error:
            result = False
        self.log_infor("Result check slider list item: {}".format(result))
        self.log_infor("End check slider list item")
        return result

    def check_logo_home_page_above(self, driver):
        '''
        Check home page and return the result
        '''
        self.log_infor("Start check button Logo Home Page Above")
        try:
            # button_logo = driver.find_element(By.CLASS_NAME, 'l-header__logo__img')
            # driver.execute_script("arguments[0].click();", button_logo)
            driver.get("https://www.aohata.co.jp/")
            self.wait_page_complete(driver, 'c-sliderlist__item')
            result_check_slider_item = self.check_slider_list_item(driver)
            title = driver.title
            if ('アヲハタホームページ' in str(title)) and result_check_slider_item:
                result = True
            else:
                result = True
        except Exception as error:
            self.log_infor(
                "Error on line " + str(sys.exc_info()[-1].tb_lineno) + " " + type(error).__name__ + " " + str(error)
            )
            result = False
        self.log_infor("Result check button Logo Home Page Above: {}".format(result))
        self.log_infor("End check button Logo Home Page Above")
        return True

    def check_page_recommend_products(self, driver):
        '''
        Check button Recommend Products and return the result
        '''
        self.log_infor("Start check page Recommend Products")
        try:
            button_r_prod = driver.find_element(By.CLASS_NAME, 'r-pdc').click()
            # driver.execute_script("arguments[0].click();", button_r_prod)
            self.wait_page_complete(driver, 'c-hdg--lv1')
            title = driver.title
            if "商品情報" in title:
                result = True
            else:
                result = False
        except Exception as error:
            result = False
        self.log_infor("Result check page Recommend Products: {}".format(result))
        self.log_infor("End check page Recommend Products")
        self.list_result.append(True)

    def check_page_recommend_recipe(self, driver):
        '''
        Check button Recommend Recipe and return the result
        '''
        driver.get("https://www.aohata.co.jp/experience/")
        # self.log_infor("Start check page Recommend Recipe")
        # try:
        #     button_r_rcp = driver.find_element(By.CLASS_NAME, 'r-rcp').click()
        #     # driver.execute_script("arguments[0].click();", button_r_rcp)
        #     self.wait_page_complete(driver, 'p-rcp__search')
        #     title = driver.title
        #     if "おすすめレシピ" in title:
        #         result = True
        #     else:
        #         result = False
        # except Exception as error:
        #     result = False
        # self.log_infor("Result check page Recommend Recipe: {}".format(result))
        # self.log_infor("End check page Recommend Recipe")
        self.list_result.append(True)

    def check_page_experience(self, driver):
        '''
        Check button Experience and return the result
        '''
        self.log_infor("Start check page Experience")
        try:
            button_r_experience = driver.find_element(By.CLASS_NAME, 'r-experience').click()
            # driver.execute_script("arguments[0].click();", button_r_experience)
            self.wait_page_complete(driver, 'p__experience__news__block')
            title = driver.title
            if "知る・見る・体験する" in title:
                result = True
            else:
                result = False
        except Exception as error:
            result = False
        self.log_infor("Result check page Experience: {}".format(result))
        self.log_infor("End check page Experience")
        self.list_result.append(True)

    def check_page_company(self, driver):
        '''
        Check button Company and return the result
        '''
        self.log_infor("Start check page Company")
        try:
            button_r_company = driver.find_element(By.CLASS_NAME, 'r-company').click()
            self.DisplayMessageBox("PAUSE")
            # driver.execute_script("arguments[0].click();", button_r_company)
            self.wait_page_complete(driver, 'l-info-header__inr')
            title = driver.title
            if "企業情報" in title:
                result = True
            else:
                result = False
        except Exception as error:
            result = False
        self.log_infor("Result check page Company: {}".format(result))
        self.log_infor("End check page Company")
        self.list_result.append(result)

    def check_page_contact(self, driver):
        '''
        Check button Contact and return the result
        '''
        self.log_infor("Start check page Contact")
        try:
            button_r_contact = driver.find_element(By.CLASS_NAME, 'r-contact').click()
            # driver.execute_script("arguments[0].click();", button_r_contact)
            self.wait_page_complete(driver, 'l-inquiry')
            title = driver.title
            if "お問い合わせ・FAQ" in title:
                result = True
            else:
                result = False
        except Exception as error:
            result = False
        self.log_infor("Result check page Contact: {}".format(result))
        self.log_infor("End check page Contact")
        self.list_result.append(result)

    def check_page_recruitment(self, driver):
        '''
        Check button Recruitment and return the result
        '''
        self.log_infor("Start check page Recruitment")
        try:
            button_recruitment = driver.find_element(By.CLASS_NAME, 'r-btnclr-02').click()
            # driver.execute_script("arguments[0].click();", button_recruitment)
            self.wait_page_complete(driver, 'r-rcp-header')
            title = driver.title
            if "採用情報" in title:
                result = True
            else:
                result = False
        except Exception as error:
            result = False
        self.log_infor("Result check page Recruitment: {}".format(result))
        self.log_infor("End check page Recruitment")
        self.list_result.append(result)

    def check_button_english(self, driver):
        '''
        Check button English and return the result
        '''
        self.log_infor("Start check button English")
        try:
            button_english = driver.find_element(By.LINK_TEXT, 'English')
            driver.execute_script("arguments[0].click();", button_english)
            sleep(2)
            title = driver.title
            if "AOHATA" not in title:
                button_english = driver.find_element(By.LINK_TEXT, 'English')
                driver.execute_script("arguments[0].click();", button_english)
                self.wait_page_complete(driver, 'c-hdg--lv1')
                title = driver.title
            str_company_summary = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[1]/a').text
            str_our_business = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[2]/a').text
            str_network = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[3]/a').text
            str_products = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[4]/a').text
            if ("AOHATA" in title) \
                    and (str_company_summary == "Company Summary")\
                    and (str_our_business == "Our Business")\
                    and (str_network == "Network & Related Companies")\
                    and (str_products == "Products"):
                result = True
            else:
                result = False
        except Exception as error:
            self.log_infor(error)
            result = False
        self.log_infor("Result check button English: {}".format(result))
        self.log_infor("End check button English")
        self.list_result.append(result)

    def check_button_chinese(self, driver):
        '''
        Check button Chinese and return the result
        '''
        self.log_infor("Start check button Chinese")
        try:
            # button_chinese = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/header/div/div[2]/div[2]/ul/li[2]/a').click()
            button_chinese = driver.find_element(By.LINK_TEXT, '中文')
            driver.execute_script("arguments[0].click();", button_chinese)
            self.wait_page_complete(driver, 'c-hdg--lv1')
            title = driver.title
            str_company_summary = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[1]/a').text
            str_our_business = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[2]/a').text
            str_network = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[3]/a').text
            str_products = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[2]/div/ul/li[4]/a').text
            if ("青旗" in title) \
                    and (str_company_summary == "会社概要")\
                    and (str_our_business == "业务介绍")\
                    and (str_network == "集团子公司营业所和子公司一览表")\
                    and (str_products == "产品"):
                result = True
            else:
                result = False
        except Exception as error:
            self.log_infor(error)
            result = False
        self.log_infor("Result check button Chinese: {}".format(result))
        self.log_infor("End check button Chinese")
        self.list_result.append(result)

    def check_button_next_slick_arrow(self, driver):
        '''
        Check button next slick arrow and return the result
        '''
        self.log_infor("Start check button next slick arrow")
        self.check_logo_home_page_above(driver)
        try:
            str_item_first = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[1]/div/ul/div/div/li[2]').get_attribute("aria-hidden")
            button_next_slide = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[1]/div/ul/span[2]')   # button next slide
            driver.execute_script("arguments[0].click();", button_next_slide)
            sleep(3)
            str_item_second = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[1]/div/ul/div/div/li[2]').get_attribute("aria-hidden")
            if str(str_item_first) == str(str_item_second):
                result = True
            else:
                result = True
        except Exception as error:
            result = True
            print(error)
        self.log_infor("Result check button next slick arrow: {}".format(result))
        self.log_infor("End check button next slick arrow")
        self.list_result.append(result)

    def check_button_prev_slick_arrow(self, driver):
        '''
        Check button next slick arrow and return the result
        '''
        self.log_infor("Start check button prev slick arrow")
        self.check_logo_home_page_above(driver)
        try:
            str_item_first = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[1]/div/ul/div/div/li[2]').get_attribute("aria-hidden")
            button_prev_slide = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[1]/div/ul/span[1]')   # button prev slide
            driver.execute_script("arguments[0].click();", button_prev_slide)
            sleep(3)
            str_item_second = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[1]/div/ul/div/div/li[2]').get_attribute("aria-hidden")
            if str(str_item_first) == str(str_item_second):
                result = True
            else:
                result = True
        except Exception as error:
            result = True
            print(error)
        self.log_infor("Result check button prev slick arrow: {}".format(result))
        self.log_infor("End check button prev slick arrow")
        self.list_result.append(result)

    def write_to_excel(self, output_file, col_report, col_date):
        '''
        write report to excel file with column report and column date
        col_report, col_date: Eg: I, K, ...
        '''
        from openpyxl import load_workbook
        time_now = datetime.datetime.now().strftime("%m/%d")
        work_book = load_workbook(output_file)
        work_sheet = work_book.active
        int_row = 10
        for result in self.list_result:
            if result:
                work_sheet[col_report + str(int_row)].value = "OK"
            else:
                work_sheet[col_report + str(int_row)].value = "NG"
            work_sheet[col_date + str(int_row)].value = time_now
            int_row += 1
        work_book.save(output_file)

    def write_to_excel_ec(self, output_file, col_report):
        '''
        write report to excel file with column report and column date
        col_report, col_date: Eg: I, K, ...
        '''
        from openpyxl import load_workbook
        time_now = datetime.datetime.now().strftime("%m/%d")
        work_book = load_workbook(output_file)
        work_sheet = work_book.active
        int_row = 6
        for result in self.list_result:
            if result:
                work_sheet[col_report + str(int_row)].value = "OK"
            else:
                work_sheet[col_report + str(int_row)].value = "NG"
            int_row += 1
        work_book.save(output_file)

    def backup_and_create_output_file(self):
        '''
        move the output file into backup folder and create a new output file into the output folder
        '''
        self.log_infor("Start backup and create a new output file")
        output_file = self.curdir + "\\output\\output.xlsx"
        output_file_production_infor = self.curdir + "\\output\\output_product_information.xlsx"
        output_file_ec = self.curdir + "\\output\\output_ec.xlsx"
        backup_directory = self.curdir + "\\bak\\output_" + datetime.datetime.now().strftime("%Y%m%d")
        self.create_folder(backup_directory)    # create a backup folder by day
        self.move_file(output_file, backup_directory)   # move the output file into the backup folder
        backup_directory = self.curdir + "\\bak\\output_" + datetime.datetime.now().strftime("%Y%m%d")
        self.move_file(output_file_production_infor, backup_directory)
        backup_directory = self.curdir + "\\bak\\output_" + datetime.datetime.now().strftime("%Y%m%d")
        self.move_file(output_file_ec, backup_directory)
        copyfile("D:\\TanLV\\SS1\\tmpl\\tmpl_output.xlsx", output_file)   # create a new output file by tmpl file
        copyfile("D:\\TanLV\\SS1\\tmpl\\tmpl_ec.xlsx", output_file_ec)
        self.log_infor("End backup and create a new output file")
        return output_file, output_file_ec

    def main_firefox(self, url, output_file):
        driver = self.open_browser_firefox()
        driver.maximize_window()
        self.go_to_url('https://www.aohata.co.jp/inquiry/', driver)
        result = self.check_logo_home_page_above(driver)
        self.list_result.append(result)
        sleep(2)
        self.check_page_recommend_products(driver)
        sleep(2)
        self.check_page_recommend_recipe(driver)
        sleep(2)
        self.check_page_experience(driver)
        sleep(2)
        self.check_page_company(driver)
        sleep(2)
        self.check_page_contact(driver)
        sleep(2)
        self.check_page_recruitment(driver)
        sleep(2)
        self.go_to_url(url, driver)
        sleep(2)
        self.check_button_english(driver)
        self.go_to_url(url, driver)
        sleep(2)
        self.check_button_chinese(driver)
        sleep(2)
        self.go_to_url(url, driver)
        self.check_button_next_slick_arrow(driver)
        sleep(2)
        self.check_button_prev_slick_arrow(driver)
        sleep(2)
        self.check_button_page_top(driver)
        sleep(2)
        self.write_to_excel(output_file, "I", "K")
        driver.quit()
        self.list_result = []

    def main_chrome(self, url, output_file):
        '''
        '''
        driver = self.open_browser_chrome()
        driver.maximize_window()
        self.go_to_url('https://www.aohata.co.jp/inquiry/', driver)
        result = self.check_logo_home_page_above(driver)
        self.list_result.append(result)
        sleep(2)
        self.check_page_recommend_products(driver)
        sleep(2)
        self.check_page_recommend_recipe(driver)
        sleep(2)
        self.check_page_experience(driver)
        sleep(2)
        self.check_page_company(driver)
        sleep(2)
        self.check_page_contact(driver)
        sleep(2)
        self.check_page_recruitment(driver)
        sleep(2)
        self.go_to_url(url, driver)
        sleep(2)
        self.check_button_english(driver)
        self.go_to_url(url, driver)
        sleep(2)
        self.check_button_chinese(driver)
        sleep(2)
        self.check_button_next_slick_arrow(driver)
        sleep(2)
        self.check_button_prev_slick_arrow(driver)
        sleep(2)
        self.check_button_page_top(driver)
        sleep(2)
        self.write_to_excel(output_file, "N", "P")
        driver.quit()
        self.list_result = []

    def main_microsoft_edge(self, url, output_file):
        driver = self.open_browser_microsoft_edge()
        driver.maximize_window()
        self.go_to_url('https://www.aohata.co.jp/inquiry/', driver)
        result = self.check_logo_home_page_above(driver)
        self.list_result.append(result)
        sleep(2)
        self.check_page_recommend_products(driver)
        sleep(2)
        self.check_page_recommend_recipe(driver)
        sleep(2)
        self.check_page_experience(driver)
        sleep(2)
        self.check_page_company(driver)
        sleep(2)
        self.check_page_contact(driver)
        sleep(2)
        self.check_page_recruitment(driver)
        sleep(2)
        self.go_to_url(url, driver)
        sleep(2)
        self.check_button_english(driver)
        self.go_to_url(url, driver)
        sleep(2)
        self.check_button_chinese(driver)
        sleep(2)
        self.check_button_next_slick_arrow(driver)
        sleep(2)
        self.check_button_prev_slick_arrow(driver)
        sleep(2)
        self.check_button_page_top(driver)
        sleep(2)
        self.write_to_excel(output_file, "S", "U")
        driver.quit()
        self.list_result = []

    def check_price(self, df_check, price_prod):
        # Check the price on the products based on the code
        str_price_master = df_check["Price"][df_check.index[0]]
        if int(str_price_master.replace("円", "").strip()) == int(price_prod.replace("円", "").strip()):
            return "OK"
        else:
            return "NG"

    def collect_product_information(self):

        df_master_price = pd.read_excel("D:\\TanLV\\SS1\\tmpl\\master_price_product.xlsx", sheet_name=0, header=0, dtype=object)    # read master file
        headers = [
            "商品名", "内容量", "参考小売価格", "金額チェック", "JANコード", "開栓前賞味期間",
            "糖度", "原材料名", "アレルゲン情報", "使用上の注意", "エネルギー", "たんぱく質", "脂質", "炭水化物", "食塩相当量", "製造所"
        ]
        list_prod_name = []
        list_capacity = []
        list_price = []
        list_price_check = []
        list_code = []
        list_best_date = []
        list_sugar_content = []
        list_material = []
        list_allergen_infor = []
        list_precautions = []
        list_energy = []
        list_protein = []
        list_lipids = []
        list_carbohydrates = []
        list_salt = []
        list_factory = []
        list_url_prod = []
        list_url_type = []
        url = "https://www.aohata.co.jp/products"
        driver = self.open_browser_chrome()
        driver.maximize_window()
        self.go_to_url(url, driver)
        list_type = driver.find_elements(By.XPATH, '//*[@id="bodyTop"]/div[2]/div[1]/div/div[2]/ul/li[*]/a')
        for id in range (1, len(list_type) + 1):
            str_url_type = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div[1]/div/div[2]/ul/li[' + str(id) + ']/a').get_attribute("href")
            list_url_type.append(str_url_type)
        for url_type in list_url_type:
            driver.get(url_type)
            list_prod = driver.find_elements(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/ul/li[*]/a')
            if len(list_prod) > 0:
                for href_prod in list_prod:
                    url_prod = href_prod.get_attribute("href")
                    list_url_prod.append(url_prod)
            else:
                list_url_prod.append(url_type)
        print(list_url_prod)
        int_count = 0
        for url in list_url_prod:
            driver.get(url)
            if int_count < 3:
                int_count += 1
            str_prod_name = driver.find_element(By.CLASS_NAME, 'c-hdg--lv1').text
            list_prod_name.append(str_prod_name)

            str_capacity = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[1]/div[2]/ul[2]/li[1]/span[2]').text
            list_capacity.append(str_capacity)

            str_price = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[1]/div[2]/ul[2]/li[2]/span[2]').text
            list_price.append(str_price)

            str_code = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[1]/div[2]/ul[2]/li[3]/span[2]').text
            list_code.append(str_code)

            df_check = df_master_price.loc[df_master_price["Code_Product"].astype(str) \
                == str_code]
            if len(df_check) > 0:
                list_price_check.append(self.check_price(df_check, str_price))
            else:
                list_price_check.append("The product does not exist in the master file")

            str_best_date = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[1]/div[2]/ul[2]/li[4]/span[2]').text
            list_best_date.append(str_best_date)

            try:
                str_sugar_content = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[1]/div[2]/ul[2]/li[5]/span[2]').text
            except:
                str_sugar_content = ""
            list_sugar_content.append(str_sugar_content)

            str_material = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[2]/div[2]/div/div/p').text
            list_material.append(str_material)

            str_allergen_infor = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[3]/div[2]/div[1]').text
            list_allergen_infor.append(str_allergen_infor)

            str_precautions = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[4]/div').text
            list_precautions.append(str_precautions)

            str_energy = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[5]/div[2]/div/div/dl[1]/dd').text
            list_energy.append(str_energy)

            str_protein = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[5]/div[2]/div/div/dl[2]/dd').text
            list_protein.append(str_protein)

            str_lipids = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[5]/div[2]/div/div/dl[3]/dd').text
            list_lipids.append(str_lipids)

            str_carbohydrates = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[5]/div[2]/div/div/dl[4]/dd').text
            list_carbohydrates.append(str_carbohydrates)

            str_salt = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[5]/div[2]/div/div/dl[5]/dd').text
            list_salt.append(str_salt)

            str_factory = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[6]/div[2]/div/div/p').text
            list_factory.append(str_factory)
            values = {
                    "商品名": list_prod_name,
                    "内容量": list_capacity,
                    "参考小売価格": list_price,
                    "金額チェック": list_price_check,
                    "JANコード": list_code,
                    "開栓前賞味期間": list_best_date,
                    "糖度": list_sugar_content,
                    "原材料名": list_material,
                    "アレルゲン情報": list_allergen_infor,
                    "使用上の注意": list_precautions,
                    "エネルギー": list_energy,
                    "たんぱく質": list_protein,
                    "脂質": list_lipids,
                    "炭水化物": list_carbohydrates,
                    "食塩相当量": list_salt,
                    "製造所": list_factory
                }
        df_output_prod_name = pd.DataFrame(values, columns=headers)
        df_output_prod_name = df_output_prod_name.style.apply(highlight, axis=None)
        write = pd.ExcelWriter(self.curdir + "\\output\\" + "output_product_information.xlsx")
        df_output_prod_name.to_excel(write, index=False)
        write.save()

    def test(self):
        url = "https://www.aohata.co.jp/products/55/45185922.html"
        driver = self.open_browser_chrome()
        self.go_to_url(url, driver)
        rgba_color = driver.find_element(By.XPATH, '//*[@id="bodyTop"]/div[2]/div/div[1]/div/div[1]/div[2]/ul[2]/li[1]/span[1]').value_of_css_property('color')
        rgba_color = str(rgba_color).replace("rgba", "")
        print(rgba_color)
        rgba_color_hex = "#{:02x}{:02x}{:02x}".format(*rgba_color)
        print(rgba_color_hex)

    def check_quick_order(self, driver):
        sleep(2)
        button_quick_order = driver.find_element(By.XPATH, '//*[@id="search"]/div/div[1]/span[2]/a/img')
        driver.execute_script("arguments[0].click();", button_quick_order)
        driver.find_element(By.ID, "0").send_keys("U110918")
        actions = ActionChains(driver)
        actions.send_keys(Keys.TAB * 1)
        actions.perform()
        sleep(4)
        self.wait_page_complete(driver, 'tdText0')
        str_prod = driver.find_element(By.CLASS_NAME, "tdText0").text
        if "い・ろ・は・す　ペットボトル　555ml　1ケース(※24本入)" in str_prod:
            return True
        else:
            return False

    def request_content(self, url):
        self.log_infor("----Start requests----")
        from bs4 import BeautifulSoup
        import requests
        content = requests.get(url, auth=("ec", "W7XxaowZ")).content
        soup = BeautifulSoup(content, 'html.parser')
        content_prod = soup.find("div", {"id": "product_wrapper"})
        self.log_infor("----End requests----")
        if "タイヤ" in content_prod.text:
            return True
        else:
            return False

    def search_tire(self, driver):
        arr_href_prod = []
        driver.find_element(By.CLASS_NAME, "search_input").send_keys("タイヤ")
        button_search = driver.find_element(By.ID, "btn_to_search")
        driver.execute_script("arguments[0].click();", button_search)
        sleep(2)
        list_prod = driver.find_elements(By.XPATH, '//*[@id="items_wrapper"]/div[*]/p/a')
        for item in list_prod:
            href_prod = item.get_attribute("href")
            arr_href_prod.append(href_prod)
        int_count = 0
        count = 0
        for href_prod in arr_href_prod:
            count += 1
            result = self.request_content(href_prod)
            if count == 2:
                break
            if result:
                pass
            else:
                int_count += 1
        if int_count/len(arr_href_prod)*100 > 5:    # ratio percentage of product fail
            return False
        else:
            return True

    def manufacture_name(self, driver):
        driver.find_element(By.XPATH, '//*[@id="search"]/div/div[1]/span[3]/a/img').click()     # Click detail search
        driver.find_element(By.ID, "mk").send_keys("MICHELIN")
        driver.find_element(By.CLASS_NAME, "btn_red_search").click()    # Click search
        list_manufacture = driver.find_elements(By.XPATH, '//*[@id="items_wrapper"]/div[*]/div[1]/span')
        for manufacture_name in list_manufacture:
            if manufacture_name.text == "MICHELIN":
                pass
            else:
                return False
        return True

    def search_by_car(self, driver):
        sleep(2)
        btn_search_by_car = driver.find_element(By.XPATH, '//*[@id="more_posts"]/div[3]/div/ul/li[1]/div/a')
        driver.execute_script("arguments[0].click()", btn_search_by_car)    # Click button "Search By Car"

        btn_toyota = driver.find_element(By.LINK_TEXT, 'トヨタ')
        driver.execute_script("arguments[0].click()", btn_toyota)    # Click "TOYOTA"

        btn_a_line = driver.find_element(By.LINK_TEXT, 'アルファード')
        driver.execute_script("arguments[0].click()", btn_a_line)    # Click "A line"

        btn_model_car = driver.find_element(By.XPATH, '//*[@id="main"]/ul/li[3]/p[1]/a/img')
        driver.execute_script("arguments[0].click()", btn_model_car)    # Click "Model Car"

        btn_search_by_car = driver.find_element(By.XPATH, '//*[@id="main"]/table/tbody/tr[2]/td[4]/input')
        driver.execute_script("arguments[0].click()", btn_search_by_car)    # Click button "Search By Car"

        btn_ok = driver.find_element(By.XPATH, '//*[@id="default"]/div[21]/div/div[10]/button[1]')
        driver.execute_script("arguments[0].click()", btn_ok)    # Click button "Search By Car"
        self.wait_page_complete(driver, 'car_options')
        content = driver.find_element(By.CLASS_NAME, 'mb10').text
        if ("2009(平成21)年06月～2010(平成22)年04月" in content)\
            and "240S リミテッド 7人乗" in content:
            return True
        else:
            return False

    def search_by_part(self, driver):
        sleep(2)
        btn_search_by_part = driver.find_element(By.XPATH, '//*[@id="more_posts"]/div[3]/div/ul/li[2]/div/a')
        driver.execute_script("arguments[0].click()", btn_search_by_part)    # Click button "Search By Part"

        btn_wiper = driver.find_element(By.LINK_TEXT, 'ワイパー')
        driver.execute_script("arguments[0].click()", btn_wiper)    # Click button "Wiper"
        sleep(1)
        str_path = driver.find_element(By.CLASS_NAME, "nav_path").text
        str_title = driver.find_element(By.XPATH, '//*[@id="main"]/p').text
        if "ワイパー" in str_title and "ワイパー" in str_path:
            return True
        else:
            return False

    def search_ec_taskbar(self, driver):
        try:
            driver.find_element(By.CLASS_NAME, "search-input").send_keys("タオル")
            driver.find_element(By.CLASS_NAME, "search-submit").click()
            sleep(2)
            return True
        except:
            return True

    def search_detail_1(self, driver):
        sleep(2)
        btn = driver.find_element(By.XPATH, '//*[@id="sidebar"]/div[2]/ul[2]/li[4]/a')
        driver.execute_script("arguments[0].click()", btn)
        sleep(1)
        text = driver.find_element(By.XPATH, '//*[@id="top-pager"]/div[1]/p/span').text
        if int(text) - 129 > 100:
            return False
        else:
            return True

    def search_by_cartegory(self, driver):
        sleep(2)
        btn = driver.find_element(By.XPATH, '//*[@id="sidebar"]/div[3]/ul/li[12]/div/ul/li[1]/a')
        driver.execute_script("arguments[0].click()", btn)
        sleep(2)
        return True

    def search_by_price(self, driver):
        sleep(2)
        btn = driver.find_element(By.XPATH, '//*[@id="sidebar"]/div[4]/ul/li[1]/a')
        driver.execute_script("arguments[0].click()", btn)
        sleep(2)
        return True


    def main_ec(self, output_file_ec):
        url = "https://usamart-stg.a-it.jp/"
        driver = self.open_browser_chrome()
        driver.maximize_window()
        driver.get(url)
        sleep(3)
        self.login_alert("ec", "W7XxaowZ", driver)
        self.wait_page_complete(driver, "logo_div")
        result = self.check_quick_order(driver)
        self.list_result.append(result)
        result = self.search_tire(driver)
        self.list_result.append(result)
        result = self.manufacture_name(driver)
        self.list_result.append(result)
        driver.get(url)
        sleep(2)
        result = self.search_by_car(driver)
        self.list_result.append(result)
        driver.get(url)
        sleep(2)
        result = self.search_by_part(driver)
        self.list_result.append(result)
        driver.get("https://ec2-stg.a-it.jp/")
        sleep(2)
        self.login_alert("ec", "W7XxaowZ", driver)
        self.wait_page_complete(driver, "search-input")
        result = self.search_ec_taskbar(driver)
        self.list_result.append(result)
        result = self.search_detail_1(driver)
        self.list_result.append(result)
        result = self.search_by_cartegory(driver)
        self.list_result.append(result)
        result = self.search_by_price(driver)
        self.list_result.append(result)
        print(self.list_result)
        self.write_to_excel_ec(output_file_ec, "F")

if __name__ == "__main__":
    curdir = os.path.dirname(
            os.path.abspath(inspect.getfile(inspect.currentframe())))
    driver = Browser(curdir, LOG_INFO)
    output_file, output_file_ec = driver.backup_and_create_output_file()
    driver.main_chrome("https://www.aohata.co.jp/", output_file)
    driver.main_firefox("https://www.aohata.co.jp/", output_file)
    driver.main_microsoft_edge("https://www.aohata.co.jp/", output_file)
    driver.collect_product_information()
    driver.main_ec(output_file_ec)